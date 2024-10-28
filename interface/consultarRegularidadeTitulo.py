from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import time
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
df = pd.DataFrame() # Dataframepara armazenamento de dados de entrada
col1 = [] # Vetores para armazenar os dados das duas colunas
ws['A1'] = "CPF"
ws['B1'] = "STATUS"

def carregar_arquivo():  # Função responsável por realizar a captura do arquivo em Excel
    global df
    # Abre uma janela para selecionar o arquivo
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])  # Método responsável por abrir a caixa de diálogo e carregar o arquivo Excel para dentro do código
    if filepath:  # Verifica se o arquivo foi selecionado
        # Carrega o arquivo Excel
        df = pd.read_excel(filepath)  # Cria o DataFrame
        
       
        # Itera sobre cada linha do DataFrame
        for i in range(len(df)):
            col1.append(df.iloc[i, 0])  # Adiciona o valor da primeira coluna ao vetor col1
    automacao() #Chama a função automação

def automacao ():
    global ws
    global df
    servico = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico)
    for i in range(len(df)):
        dados = str(df[i])
        navegador.get("https://www.tse.jus.br/servicos-eleitorais/autoatendimento-eleitoral#/atendimento-eleitor/consultar-situacao-titulo-eleitor")# Passo 1: Acessar o site
        time.sleep(2)
        navegador.find_element('xpath', '//*[@id="modal-lgpd"]/div/div/div[2]/button').click()
        time.sleep(2)
        navegador.find_element('xpath', '//*[@id="titulo-cpf-nome"]').send_keys(dados)# Passo 2: Inserir o CPF
        navegador.find_element('xpath', '//*[@id="modal"]/div/div/div[2]/div[2]/form/div[2]/button[2]').click() # Passo 3: Clicar no botão de consulta
        cpf = navegador.find_element('xpath','//*[@id="content"]/app-root/div/app-consultar-situacao-titulo-eleitor/app-avatar-e-nome/div/div/div/div[2]/p/span').text#Capturar o CPF e o Status
        status = navegador.find_element('xpath','//*[@id="content"]/app-root/div/app-consultar-situacao-titulo-eleitor/div[1]/div[1]/p/span').text
        ws.append([cpf,status])
        if i < len(df) - 1:  
            ws.save("c:/PROGRAMACAO/PYTHON/excel/teste.xlsx")

# Cria a janela principal
root = tk.Tk()
root.title("Carregar Arquivo Excel")
# Define o tamanho da janela (largura x altura)
root.geometry("300x100")
# Cria um botão para carregar o arquivo
btn_carregar = tk.Button(root, text="Carregar Arquivo Excel", command=carregar_arquivo)
btn_carregar.pack(pady=20)
# Inicia o loop da interface gráfica
root.mainloop()
