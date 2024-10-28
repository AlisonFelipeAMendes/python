from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = "CPF"
ws['B1'] = "STATUS"

numero = int(input("Quantos CPF teram na tabela: "))


for i in range(1,numero):
    ws.append([i,i+2])
#ws.append([3,4])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("c:/PROGRAMACAO/PYTHON/excel/sample.xlsx")