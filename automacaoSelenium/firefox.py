from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
import time
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
df = pd.DataFrame()  # DataFrame para armazenamento de dados de entrada
col1 = []  # Vetores para armazenar os dados das duas colunas
ws['A1'] = "CPF"
ws['B1'] = "STATUS"

def carregar_arquivo():
    global df
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if filepath:
        df = pd.read_excel(filepath)
        for i in range(len(df)):
            col1.append(df.iloc[i, 0])  # Adiciona o valor da primeira coluna ao vetor col1
    automacao()  # Chama a função automação

def automacao():
    global ws
    global df
    options = Options()
    options.headless = True  # Executa o navegador em modo headless
    servico = Service(GeckoDriverManager().install())
    navegador = webdriver.Firefox(service=servico, options=options)
    
    for i in range(len(df)):
        dados = str(df.iloc[i, 0])  # Correção aqui: Use df.iloc[i, 0] para acessar o valor da célula
        navegador.get("https://www.tse.jus.br/servicos-eleitorais/autoatendimento-eleitoral#/atendimento-eleitor/consultar-situacao-titulo-eleitor")
        
        try:
            # Aguarda até que o botão esteja clicável e então clica
            botao_lgpd = WebDriverWait(navegador, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="modal-lgpd"]/div/div/div[2]/button'))
            )
            botao_lgpd.click()
            
            # Continua para a próxima etapa
            campo_cpf = WebDriverWait(navegador, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="titulo-cpf-nome"]'))
            )
            campo_cpf.send_keys(dados)
            
            # Clica no botão de consulta
            botao_consulta = navegador.find_element(By.XPATH, '//*[@id="modal"]/div/div/div[2]/div[2]/form/div[2]/button[2]')
            botao_consulta.click()
            
            time.sleep(2)  # Aguarda a resposta
            
            # Captura os valores de CPF e Status
            cpf = navegador.find_element(By.XPATH, '//*[@id="content"]/app-root/div/app-consultar-situacao-titulo-eleitor/app-avatar-e-nome/div/div/div/div[2]/p/span').text
            status = navegador.find_element(By.XPATH, '//*[@id="content"]/app-root/div/app-consultar-situacao-titulo-eleitor/div[1]/div[1]/p/span').text
            ws.append([cpf, status])
        
        except Exception as e:
            print(f"Ocorreu um erro: {e}")
        
        finally:
            navegador.quit()  # Fecha o navegador
            
    # Salva o workbook no final do processo
    wb.save("c:/PROGRAMACAO/PYTHON/excel/teste.xlsx")

# Cria a janela principal
root = tk.Tk()
root.title("Carregar Arquivo Excel")
root.geometry("300x100")
btn_carregar = tk.Button(root, text="Carregar Arquivo Excel", command=carregar_arquivo)
btn_carregar.pack(pady=20)
root.mainloop()

# Salva o workbook no final do processo
wb.save("c:/PROGRAMACAO/PYTHON/excel/teste.xlsx")
